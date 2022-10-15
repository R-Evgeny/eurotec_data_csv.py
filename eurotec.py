import openpyxl
import pymysql
import bs4
import requests
from config import host, user, password, db_name

# eurotec_sqp = []
# eurotec_sq = []
#
# #Формирование БСХ
# book = openpyxl.open('bsh.xlsx', read_only=True)
# sheet = book.active
# for row in range(3, sheet.max_row+1):
#     sku = sheet[row][0].value
#     quantity = sheet[row][6].value
#     for x, y in ('Больше 5', '5'), ('Есть', '2'), ('Ожидается', '0'), ('Нет', '0'):
#         quantity = quantity.replace(x, y)
#     bsh = []
#     bsh.append(quantity)
#     bsh.append(sku)
#     eurotec_sq.append(bsh)
# #print(eurotec_sq)
# print("BSH - ОК")
#
# #Формирование Teka
# book = openpyxl.open('teka.xlsx', read_only=True)
# sheet = book.active
# for row in range(20, sheet.max_row+1):
#     sku = sheet[row][0].value
#     quantity = str(sheet[row][2].value)
#     pr = sheet[row][3].value
#     quantity = quantity.replace('> 10', '10')
#     teka = []
#     teka.append(quantity)
#     teka.append(pr)
#     teka.append(sku)
#     eurotec_sqp.append(teka)
# #print(eurotec_sqp)
# print("TEKA - ОК")
#
# #Запись Franke в CSV
# book = openpyxl.open('franke.xlsx', read_only=True)
# sheet = book.active
# for row in range(2, sheet.max_row+1):
#     sku = sheet[row][1].value
#     quantity = str(sheet[row][3].value)
#     quantity = quantity.replace('наявність по запиту', '0')
#     franke = []
#     franke.append(quantity)
#     franke.append(sku)
#     eurotec_sq.append(franke)
# #print(eurotec_sq)
# print("Franke - ОК")
#
# #Запись Mirs в CSV
#
# book = openpyxl.open('mirs.xlsx', read_only=True)
# sheet = book.active
# for row in range(7, sheet.max_row+1):
#     manufacturer = sheet[row][1].value
#     sku = sheet[row][2].value
#     quantity = str(sheet[row][8].value)
#     quantity = quantity.replace('10+', '10')
#     price = sheet[row][9].value
#     mirs = []
#     if quantity != '0':
#         if manufacturer == 'Blanco':
#             mirs.append(int(quantity))
#             mirs.append(sku)
#             mirs.append(price)
#         elif manufacturer == 'Falmec':
#             mirs.append(int(quantity))
#             mirs.append(sku)
#             mirs.append(price)
#         elif manufacturer == 'Liebherr':
#             mirs.append(int(quantity))
#             mirs.append(sku)
#             mirs.append(price)
#         elif manufacturer == 'Nivona':
#             mirs.append(int(quantity))
#             mirs.append(sku)
#             mirs.append(price)
#         else:
#             continue
#     else:
#         continue
#     eurotec_sqp.append(mirs)
# #print(eurotec_sqp)
# print("Mirs - ОК")
#
# print(eurotec_sq)
# print(eurotec_sqp)

try:
    connection = pymysql.connect(
        host=host,
        port=3306,
        user=user,
        password=password,
        database=db_name,
        cursorclass=pymysql.cursors.DictCursor
    )
    print("successfully connected...")
    print("#" * 20)

    # try:
    #     with connection.cursor() as cursor:
    #         update_000 = "UPDATE `oc_product` SET `quantity`= 0 WHERE `mpn` = 'TG'"
    #         update_query_sqp = "UPDATE oc_product SET quantity = %s, price = %s WHERE sku = %s"
    #         update_query_sq = "UPDATE oc_product SET quantity = %s WHERE sku = %s"
    #         cursor.execute(update_000)
    #         print('Quantity = 000')
    #         cursor.executemany(update_query_sqp, mymagaz_sqp)
    #         print('sqp - OK')
    #         cursor.executemany(update_query_sq, mymagaz_sq)
    #         print('sq - OK')
    #         connection.commit()
    #         print('All Updated')
    #
    # except Exception as ex:
    #     print("Error...")
    #     print(ex)
    #
    # finally:
    #     connection.close()
    #     print("Connection close")
except Exception as ex:
    print("Connection refused...")
    print(ex)

finally:
    print("All Complete")
