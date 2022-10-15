import openpyxl
import csv
import datetime

starttime = datetime.datetime.now()
starttime1 = datetime.datetime.now().strftime('%d_%m_%Y_%H_%M')
with open(f"eurotec_data_{starttime1}.csv", "w", newline='', encoding="cp1251") as file:
    writer = csv.writer(file)
    writer.writerow(
        (
            "sku",
            "quantity"
        )
    )

# Формирование БСХ
book = openpyxl.open('bsh.xlsx', read_only=True)
sheet = book.active
for row in range(13, sheet.max_row+1):
    sku = sheet[row][4].value
    quantity = sheet[row][7].value
    quantity = quantity.replace("да", "2")
    if quantity == 'нет':
        continue
    else:
        with open(f"eurotec_data_{starttime1}.csv", "a", newline='', encoding="cp1251") as file:
            writer = csv.writer(file)
            writer.writerow(
                (
                    sku,
                    quantity
                )
            )
print("BSH - ОК")

# Формирование Teka
book = openpyxl.open('teka.xlsx', read_only=True)
sheet = book.active
for row in range(17, sheet.max_row+1):
    sku = str(sheet[row][0].value)
    quantity = str(sheet[row][2].value)
    quantity = quantity.replace('> 10', '10')
    with open(f"eurotec_data_{starttime1}.csv", "a", newline='', encoding="cp1251") as file:
        writer = csv.writer(file)
        writer.writerow(
            (
                sku,
                quantity
            )
        )
print("TEKA - ОК")

# #Формирование Franke
book = openpyxl.open('franke.xlsx', read_only=True)
sheet = book.active
for row in range(2, sheet.max_row+1):
    sku = sheet[row][1].value
    quantity = str(sheet[row][3].value)
    if quantity == 'наявність по запиту':
        continue
    else:
        with open(f"eurotec_data_{starttime1}.csv", "a", newline='', encoding="cp1251") as file:
            writer = csv.writer(file)
            writer.writerow(
                (
                    sku,
                    quantity
                )
            )
print("Franke - ОК")

#Формирование Mirs
book = openpyxl.open('mirs.xlsx', read_only=True)
sheet = book.active
mirs_all = []
for row in range(7, sheet.max_row+1):
    manufacturer = sheet[row][1].value
    sku = sheet[row][2].value
    quantity = str(sheet[row][8].value)
    quantity = quantity.replace('10+', '10')
    mirs = []
    if quantity != '0':
        if manufacturer == 'Blanco':
            with open(f"eurotec_data_{starttime1}.csv", "a", newline='', encoding="cp1251") as file:
                writer = csv.writer(file)
                writer.writerow(
                    (
                        sku,
                        quantity
                    )
                )
        elif manufacturer == 'Falmec':
            with open(f"eurotec_data_{starttime1}.csv", "a", newline='', encoding="cp1251") as file:
                writer = csv.writer(file)
                writer.writerow(
                    (
                        sku,
                        quantity
                    )
                )
        elif manufacturer == 'Liebherr':
            with open(f"eurotec_data_{starttime1}.csv", "a", newline='', encoding="cp1251") as file:
                writer = csv.writer(file)
                writer.writerow(
                    (
                        sku,
                        quantity
                    )
                )
        elif manufacturer == 'Nivona':
            with open(f"eurotec_data_{starttime1}.csv", "a", newline='', encoding="cp1251") as file:
                writer = csv.writer(file)
                writer.writerow(
                    (
                        sku,
                        quantity
                    )
                )
        elif manufacturer == 'Vestel':
            with open(f"eurotec_data_{starttime1}.csv", "a", newline='', encoding="cp1251") as file:
                writer = csv.writer(file)
                writer.writerow(
                    (
                        sku,
                        quantity
                    )
                )
        else:
            continue
    else:
        continue
print("Mirs - ОК")
diftime = datetime.datetime.now() - starttime
print(starttime)
print(diftime)
